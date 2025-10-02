import openpyxl
from config import Data_file

## ToDo
## please add "_" to private classes
## add descriptions for the other methods

#gets and returns file path of excel file
def initiate_file():
    data_file = Data_file()
    file_path = data_file.load_data_file_path()
    return file_path

#reads all consecutive not-empty cells in a row
def read_columns_in_row(sheet, row) -> list:
    values = []
    for column in range(1, 100):
        cell = sheet.cell(row=row, column=column)
        if(cell.value is not None):
            values.append(cell.value)
        else:
            break
    return values

# reads all data of one class from excel file
def get_class_data(sheet, title_row):
    class_title = sheet.cell(row=title_row, column=1).value
    headers = read_columns_in_row(sheet=sheet, row=title_row + 3) # wir schauen hier nochmal in den großen Listen, was der Abstand ist, 3 Zielen im Testding
    print(class_title)
    for header in headers:
        print(header)

def read_excel_file(sheet, start_row):
    row_count = sheet.max_row
    for row in range(start_row, row_count):
        cell = sheet.cell(row=row, column=1)

        if cell.value is not None:
            get_class_data(sheet=sheet, title_row=row)
            break;

def run():
    file_path = initiate_file()
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    read_excel_file(sheet, start_row=1)

run()
