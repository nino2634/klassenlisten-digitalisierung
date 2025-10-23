from config_handler import Config_Handler
import openpyxl

def init_file():
    data_file = Config_Handler()
    file_path = data_file.load_data_file_path()
    return file_path

def get_classes_from_file(sheet, start_row, class_filter): 
    classes = []
    spacer = 3 # empty rows between class name and lessons
    end_row = sheet.max_row
    row = start_row
    while end_row > row:
        cell = sheet.cell(row=row, column=1)
        if cell.value is not None:
            print(cell.value)
            if(class_filter in cell.value):
                classes.append(cell.value)
            if(row+spacer < end_row):
                new_row=skip_lessons(sheet=sheet, start_row=row+spacer, end_row=end_row)
                row = new_row
            else:
                break;
        else:
            row += 1
    return classes

def skip_lessons(sheet, start_row, end_row):
    spacer = 1 # empty rows between lessons and next class name
    for row in range(start_row, end_row):
        print(row)
        cell = sheet.cell(row=row, column=1)
        print(cell.value)
        if(cell.value is None and row != end_row):
            return (row + spacer)
    return end_row

def run(class_filter = ""):
    print(class_filter)
    filePath = init_file()
    workbook = openpyxl.load_workbook(filePath)
    sheet=workbook.active
    classes = get_classes_from_file(sheet=sheet, start_row=1, class_filter=class_filter)
    print(classes)

run()