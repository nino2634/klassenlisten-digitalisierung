# -*- coding: utf-8 -*-
import json
import openpyxl
from .get_headers import run as get_headers
from .file_handler import initiate_file, get_next_empty_row, get_next_row_with_value
from .config.config_handler import load_config_data

# variables
classes_column_name = load_config_data("classes_column_name")
weekly_hrs_column_name = load_config_data("weekly_hrs_column_name")
half_year_column_name = load_config_data("half_year_column_name")#.encode("latin-1").decode("utf-8")
subject_column_name = load_config_data("subject_column_name")
teacher_column_name = load_config_data("teacher_column_name")

def _find_class_title_row(sheet, class_filter):
    end_row = sheet.max_row
    for row in range(1, end_row):
        cell = sheet.cell(row=row, column=sheet.min_column)
        if cell.value == class_filter:
            return row
    raise Exception("Class not found")
    
def _get_lessons_by_class(sheet, class_title, con_year_half, headers):
    title_row = _find_class_title_row(sheet, class_title)
    start_row = get_next_row_with_value(sheet, title_row + 1) + 1
    end_row = get_next_empty_row(sheet, start_row)

    start_col = sheet.min_column
    end_col = sheet.max_column
    subject_column = headers.index(subject_column_name) + 1

    lessons_list = []
    sum_sus = 0
    sum_kuk = 0

    row = start_row
    while row < end_row:
        lesson = {}
        mergeRows = False

        # nur Zeilen berücksichtigen, die nicht zum aktuellen Halbjahr gehören
        if con_year_half not in str(sheet.cell(row, headers.index(half_year_column_name) + 1).value):
            # Werte aus der Zeile auslesen
            for col in range(start_col, end_col):
                cell = sheet.cell(row, col)
                header = headers[col - 1]
                lesson[header] = f"{cell.value}"

                if weekly_hrs_column_name in header and (cell.value is None or cell.value == 'None'):
                    lesson[header] = 0

                if headers[col - 1] == subject_column_name:
                    next_cell = sheet.cell(row + 1, col)
                    if cell.value == next_cell.value:
                        mergeRows = True

            # Kommentar für gekoppelte Klassen
            if "," in lesson[classes_column_name]:
                classes = [c.strip() for c in lesson[classes_column_name].split(",")]
                comment = "gekoppelt mit"
                for class_name in classes:
                    if class_name != class_title:
                        comment += f" {class_name},"
                lesson['comment'] = comment.rstrip(",")

            # Merge-Logik
            nextRow = row + 1
            while mergeRows and nextRow < end_row:
                """# SuS Stunden addieren
                val = sheet.cell(nextRow, headers.index(f"{weekly_hrs_column_name}_SuS") + 1).value
                if val is not None and val != 'None':
                    lesson[f"{weekly_hrs_column_name}_SuS"] = float(lesson[f"{weekly_hrs_column_name}_SuS"]) + float(val)

                # KuK Stunden addieren
                val = sheet.cell(nextRow, headers.index(f"{weekly_hrs_column_name}_KuK") + 1).value
                if val is not None and val != 'None':
                    lesson[f"{weekly_hrs_column_name}_KuK"] = float(lesson[f"{weekly_hrs_column_name}_KuK"]) + float(val)
                """
                # Lehrer zusammenführen
                if sheet.cell(nextRow, headers.index(teacher_column_name) + 1).value not in lesson[teacher_column_name]:
                    lesson[teacher_column_name] += f", {sheet.cell(nextRow, headers.index(teacher_column_name) + 1).value}"

                # Nächste Zeile
                nextRow += 1

                # Abbruchbedingung: Ende erreicht oder Subject unterschiedlich
                if nextRow >= end_row or sheet.cell(nextRow, subject_column).value != sheet.cell(row, subject_column).value:
                    mergeRows = False

            # row auf nextRow setzen, damit äußere Schleife korrekt weiterläuft
            row = nextRow

            # Lesson speichern
            lessons_list.append(lesson)

            # Summen berechnen
            sum_sus += float(lesson[f"{weekly_hrs_column_name}_SuS"])
            sum_kuk += float(lesson[f"{weekly_hrs_column_name}_KuK"])
        else:
            # Halbjahr passt nicht, einfach zur nächsten Zeile
            row += 1

    # Runden
    sum_sus = round(sum_sus, 2)
    sum_kuk = round(sum_kuk, 2)

    return {
        "class_name": f"{class_title}",
        "lessons": lessons_list,
        "Sum_SuS": sum_sus,
        "Sum_KuK": sum_kuk
    }

def run(class_title, year_half):
    con_year_half = year_half
    if "1.Hj" in year_half:
        con_year_half = str.replace(con_year_half, "1.Hj", "2.Hj")
    elif "2.Hj" in year_half:
        con_year_half = str.replace(con_year_half, "2.Hj", "1.Hj")
    
    workbook = initiate_file()
    sheet = workbook.active
    headers = json.loads(get_headers())
    lesson_list = _get_lessons_by_class(sheet, class_title, con_year_half, headers)  
    lesson_json = json.dumps(lesson_list)  
    return lesson_json

result = run("02TSBR", "1.Hj")
#print(result)