import openpyxl
import json
from .file_handler import initiate_file, get_next_row_with_value

# searches all class names in excel file
# if filter is set, searches classes with specific names
# returns: class names as list
def _get_classes_from_file(sheet, start_row, class_filter): 
    classes = []
    spacer = 0
    end_row = sheet.max_row
    row = start_row
    while end_row > row:
        cell = sheet.cell(row=row, column=sheet.min_column)
        if cell.value is not None:
            if class_filter in cell.value:
                classes.append(cell.value)
                spacer = 1 + get_next_row_with_value(sheet, start_row=row + 1) - row

            if row+spacer < end_row:
                new_row=_skip_lessons(sheet=sheet, start_row=row+spacer, end_row=end_row)
                row = new_row
            else:
                break;
        else:
            row += 1
    return classes

# skips all lessons to get to the next class name
# returns: row where list of lessons end as Int
def _skip_lessons(sheet, start_row, end_row):
    for row in range(start_row, end_row):
        cell = sheet.cell(row=row, column=sheet.min_column)
        if cell.value is None and row < end_row:
            return (row)
    return end_row

# executes functions
# returns class names as list
def run(class_filter = ""):
    workbook = initiate_file()
    sheet=workbook.active
    classes = _get_classes_from_file(sheet=sheet, start_row=1, class_filter=class_filter)
    if classes == []:
        errormsg = f"Class '{class_filter}' not found"
        return json.dumps(errormsg)
    classes_json = json.dumps(classes)
    return classes_json

run()