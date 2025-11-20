import openpyxl
import json
from .file_handler import initiate_file
from .config.config_handler import load_config_data

# reads all consecutive not-empty cells in a row
# returns: column names as list
def _read_columns_in_row(sheet, row) -> list:
    values = []
    weekly_hrs_header = load_config_data("weekly_hrs_column_name")
    for column in range(sheet.min_column, sheet.max_column):
        cell = sheet.cell(row=row, column=column)
        upperCell = ""
        if(cell.value is not None):
            if cell.value == weekly_hrs_header:
                upperCell = "_"+sheet.cell(row=row-int(load_config_data("row_diff_weekly_hrs")), column=column).value # add _SuS or _KuK to WoStd
            values.append(cell.value+upperCell)
        else:
            break
    return values

# reads all headers from excel file
# returns: header names as list
def _get_headers(sheet, title_row):
    headers = _read_columns_in_row(sheet=sheet, row=title_row + 3) # wir schauen hier nochmal in den großen Listen, was der Abstand ist, 3 Zielen im Testding
    list = []
    for header in headers:
        #print(header)
        list.append(header)
    return list

# gets first row with value out of the excel file
# returns: list
def _read_excel_file(sheet, start_row): 
    row_count = sheet.max_row
    for row in range(start_row, row_count):
        cell = sheet.cell(row=row, column=sheet.min_column)
        if cell.value is not None:
            headers = _get_headers(sheet=sheet, title_row=row)
            return headers

# executes functions
# returns: header names as list
def run():
    workbook = initiate_file()
    sheet = workbook.active
    headers = _read_excel_file(sheet, start_row=1)
    headers_json = json.dumps(headers)
    #print(headers)
    return headers_json

#print(run())