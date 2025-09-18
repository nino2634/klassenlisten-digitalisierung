import openpyxl
from app.core.config import Data_file

#file_path = "/home/niklas/Coding/klassenlisten-digitalisierung/src/app/data/Datenquelle.xlsx"
df = data_file()
file_path = df.load_data_file_path()

def read_excel_file():
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    row_count = sheet.max_row

    for row in range(row_count):
        if row > 0:
            cell=sheet.cell(row=row, column=1)

            print(cell.value)

read_excel_file()
