from .config.config_handler import load_config_data
import openpyxl

def initiate_file():
    """
    Öffnet die Excel-Datei und gibt das Workbook-Objekt zurück.
    Returns: geöffnetes Workbook als openpyxl-Workbook-Objekt
    """
    filePath = load_config_data("excel_file")
    workbook = openpyxl.load_workbook(filePath)
    return workbook

def get_next_empty_row(sheet, start_row):
    """
    Sucht in gegebenem Sheet die nächste leere Zeile ab start_row in der ersten Spalte.
    
    :param sheet: openpyxl Worksheet Objekt
    :param start_row: integer, Startzeile für die Suche
    
    Returns: integer, Zeilennummer der nächsten leeren Zeile
    """
    end_row = sheet.max_row # == letze Zeile der Datei mit Inhalt
    for row in range(start_row, end_row):
        cell = sheet.cell(row=row, column=sheet.min_column)
        if(cell.value is None):
            return row
        
def get_next_row_with_value(sheet, start_row):
    """
    Sucht in gegebenem Sheet die nächste Zeile mit Inhalt ab start_row in der ersten Spalte.
    
    :param sheet: openpyxl Worksheet Objekt
    :param start_row: integer, Startzeile für die Suche
    """
    end_row = sheet.max_row # == letze Zeile der Datei mit Inhalt
    for row in range(start_row, end_row):
        cell=sheet.cell(row=row, column=sheet.min_column)
        if(cell.value is not None):
            return row
